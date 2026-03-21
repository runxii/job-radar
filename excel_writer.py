"""
Stage 5 - Excel Writer
Maintains a local jobs.xlsx with three sheets: raw / matched / unmatched.
Implements upsert-by-id semantics, mirroring the original Google Sheets nodes.
"""
from __future__ import annotations
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import config
 
# Column definitions per sheet
RAW_COLS = [
    "id", "title", "company", "post_url", "description", "location",
]
MATCHED_COLS = [
    "id", "title", "company","post_url", "description", "location",
    "explicit_years_required",
    "match_score", "status",
]
UNMATCHED_COLS = [
    "id", "title", "company", "location",
    "explicit_years_required", "exp_evidence",
    "is_explicit_exp_requirement",  "post_url", "description",
]
 
_HEADER_FILL   = PatternFill("solid", start_color="2F5496", end_color="2F5496")
_HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_CELL_FONT     = Font(name="Arial", size=11)
_FROZEN_ROW    = "A2"
_COL_WIDTHS    = {"description": 60, "exp_evidence": 40}
_DEFAULT_WIDTH = 20
 
 
# --------------------------------------------------------------------------- #
# Internal helpers                                                              #
# --------------------------------------------------------------------------- #
 
def _ensure_workbook(path: str) -> None:
    """Create workbook with correct sheets if it does not exist."""
    if os.path.exists(path):
        return
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, cols in [
        (config.SHEET_RAW,       RAW_COLS),
        (config.SHEET_MATCHED,   MATCHED_COLS),
        (config.SHEET_UNMATCHED, UNMATCHED_COLS),
    ]:
        ws = wb.create_sheet(sheet_name)
        _write_header(ws, cols)
    wb.save(path)
    wb.close()
 
 
def _write_header(ws, cols: list[str]) -> None:
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        width = _COL_WIDTHS.get(col_name, _DEFAULT_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = _FROZEN_ROW
 
 
def _load_id_to_row(ws) -> dict[str, int]:
    """Return {id_value: row_number} mapping from an existing sheet."""
    try:
        id_col = [cell.value for cell in ws[1]].index("id") + 1
    except ValueError:
        return {}
    return {
        ws.cell(row=r, column=id_col).value: r
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=id_col).value is not None
    }
 
 
def _upsert_rows(ws, records: list[dict], cols: list[str]) -> None:
    """Insert new rows or update existing ones (matched by 'id')."""
    id_to_row = _load_id_to_row(ws)
 
    for record in records:
        row_id = str(record.get("id", ""))
        if row_id in id_to_row:
            row_num = id_to_row[row_id]
        else:
            row_num = ws.max_row + 1
            id_to_row[row_id] = row_num
 
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=record.get(col_name, ""))
            cell.font = _CELL_FONT
            cell.alignment = Alignment(wrap_text=False)
 
 
# --------------------------------------------------------------------------- #
# Public API                                                                   #
# --------------------------------------------------------------------------- #
 
def write_raw(records: list[dict], path: str = config.OUTPUT_EXCEL) -> None:
    _ensure_workbook(path)
    wb = load_workbook(path)
    try:
        _upsert_rows(wb[config.SHEET_RAW], records, RAW_COLS)
        wb.save(path)
    finally:
        wb.close()
    print(f"[excel] raw sheet updated - {len(records)} records")
 
 
def write_matched(records: list[dict], path: str = config.OUTPUT_EXCEL) -> None:
    _ensure_workbook(path)
    wb = load_workbook(path)
    try:
        _upsert_rows(wb[config.SHEET_MATCHED], records, MATCHED_COLS)
        wb.save(path)
    finally:
        wb.close()
    print(f"[excel] matched sheet updated - {len(records)} records")
 
 
def write_unmatched(records: list[dict], path: str = config.OUTPUT_EXCEL) -> None:
    _ensure_workbook(path)
    wb = load_workbook(path)
    try:
        _upsert_rows(wb[config.SHEET_UNMATCHED], records, UNMATCHED_COLS)
        wb.save(path)
    finally:
        wb.close()
    print(f"[excel] unmatched sheet updated - {len(records)} records")
 
 
def read_matched_ids(path: str = config.OUTPUT_EXCEL) -> set[str]:
    """Return set of job ids already in the matched sheet (for dedup across runs)."""
    if not os.path.exists(path):
        return set()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if config.SHEET_MATCHED not in wb.sheetnames:
            return set()
        ws = wb[config.SHEET_MATCHED]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return set()
        try:
            id_col = list(rows[0]).index("id")
        except ValueError:
            return set()
        return {str(row[id_col]) for row in rows[1:] if row[id_col] is not None}
    finally:
        wb.close()
 