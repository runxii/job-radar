import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest
import tempfile
from openpyxl import load_workbook
import config
from excel_writer import (
    write_raw, write_matched, write_unmatched, read_matched_ids,
    RAW_COLS, MATCHED_COLS, UNMATCHED_COLS,
)


def _tmp_path(tmp_dir):
    return os.path.join(tmp_dir, "test_jobs.xlsx")


def _raw_job(job_id, title="Dev"):
    return {
        "id": job_id, "title": title, "company": "ACME",
        "posted_at": "2024-01-01", "apply_url": "https://example.com",
        "applicants": "N/A", "post_url": "https://linkedin.com/jobs/view/1",
        "description": "A job", "location": "Dublin",
    }


def _matched_job(job_id):
    job = _raw_job(job_id)
    job.update({
        "explicit_years_required": 2, "is_explicit_exp_requirement": True,
        "exp_evidence": "2 years exp", "match_score": 0.75,
        "skills_required": '{"stack": 0.8}', "status": "AI Apply",
    })
    return job


class TestWriteRaw:
    def test_creates_file_and_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _tmp_path(tmp)
            write_raw([_raw_job("1")], path)
            assert os.path.exists(path)
            wb = load_workbook(path)
            assert config.SHEET_RAW in wb.sheetnames
            assert config.SHEET_MATCHED in wb.sheetnames
            assert config.SHEET_UNMATCHED in wb.sheetnames

    def test_header_row_correct(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _tmp_path(tmp)
            write_raw([_raw_job("1")], path)
            ws = load_workbook(path)[config.SHEET_RAW]
            headers = [ws.cell(row=1, column=i+1).value for i in range(len(RAW_COLS))]
            assert headers == RAW_COLS

    def test_data_written(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _tmp_path(tmp)
            write_raw([_raw_job("42", title="SWE")], path)
            ws = load_workbook(path)[config.SHEET_RAW]
            id_col   = RAW_COLS.index("id") + 1
            title_col = RAW_COLS.index("title") + 1
            assert ws.cell(row=2, column=id_col).value   == "42"
            assert ws.cell(row=2, column=title_col).value == "SWE"


class TestUpsertBehaviour:
    def test_no_duplicate_on_second_write(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _tmp_path(tmp)
            write_raw([_raw_job("1")], path)
            write_raw([_raw_job("1")], path)   # same id again
            ws = load_workbook(path)[config.SHEET_RAW]
            # Row 1 = header; only one data row expected
            data_rows = [r for r in range(2, ws.max_row + 1)
                         if ws.cell(row=r, column=RAW_COLS.index("id") + 1).value is not None]
            assert len(data_rows) == 1

    def test_update_existing_record(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _tmp_path(tmp)
            write_raw([_raw_job("1", "OldTitle")], path)
            write_raw([_raw_job("1", "NewTitle")], path)
            ws = load_workbook(path)[config.SHEET_RAW]
            title_col = RAW_COLS.index("title") + 1
            assert ws.cell(row=2, column=title_col).value == "NewTitle"

    def test_new_record_appended(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _tmp_path(tmp)
            write_raw([_raw_job("1")], path)
            write_raw([_raw_job("2")], path)
            ws = load_workbook(path)[config.SHEET_RAW]
            id_col = RAW_COLS.index("id") + 1
            ids = [ws.cell(row=r, column=id_col).value for r in range(2, ws.max_row + 1)]
            assert "1" in ids and "2" in ids


class TestReadMatchedIds:
    def test_returns_empty_set_when_no_file(self):
        ids = read_matched_ids("/tmp/nonexistent_jobs_xyz.xlsx")
        assert ids == set()

    def test_returns_correct_ids(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _tmp_path(tmp)
            write_matched([_matched_job("10"), _matched_job("20")], path)
            ids = read_matched_ids(path)
            assert "10" in ids
            assert "20" in ids

    def test_excludes_ids_not_in_matched(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _tmp_path(tmp)
            write_raw([_raw_job("99")], path)
            write_matched([_matched_job("10")], path)
            ids = read_matched_ids(path)
            assert "99" not in ids


class TestWriteUnmatched:
    def test_writes_unmatched_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _tmp_path(tmp)
            job = _raw_job("5")
            job.update({"explicit_years_required": 5, "is_explicit_exp_requirement": True,
                        "exp_evidence": "5 years", })
            write_unmatched([job], path)
            ws = load_workbook(path)[config.SHEET_UNMATCHED]
            id_col = UNMATCHED_COLS.index("id") + 1
            assert ws.cell(row=2, column=id_col).value == "5"
